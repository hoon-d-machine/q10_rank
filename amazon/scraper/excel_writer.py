from pathlib import Path

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

HEADERS = [
    "timestamp",
    "category_key",
    "category_label",
    "rank",
    "asin",
    "name",
    "price",
    "rating",
    "review_count",
    "url",
]


def _new_workbook():
    wb = Workbook()
    ws = wb.active
    ws.title = "rankings"
    ws.append(HEADERS)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}1"
    return wb, ws


def append_rows(excel_path, rows):
    """rows: dict 리스트. 기존 파일이 있으면 append, 없으면 새로 생성.
    누적 방식이므로 기존 데이터는 절대 덮어쓰지 않는다."""
    path = Path(excel_path)
    path.parent.mkdir(parents=True, exist_ok=True)

    if path.exists():
        wb = openpyxl.load_workbook(path)
        ws = wb["rankings"] if "rankings" in wb.sheetnames else wb.active
    else:
        wb, ws = _new_workbook()

    for row in rows:
        ws.append([row.get(h) for h in HEADERS])

    wb.save(path)
